from time import sleep
from datetime import time
import openpyxl
from  openpyxl.utils import get_column_letter
from openpyxl.styles import Font , Fill , Alignment
from db import DataBase
import excel2img

class Xmlwork:
    def __init__(self,data):
        self.wb = openpyxl.Workbook()
        self.wb.create_sheet(title='Первый лист', index=0)
        self.sheet = self.wb['Первый лист']

        self.fill_data(data)
        self.setting_up()

    def fill_data(self,data):
        self.cur_row = 1
        for row in range(1, 10):
            if self.cur_row == 1:
                cell = self.sheet.cell(row=self.cur_row, column=1)
                cell.value = data["day"]
                cell = self.sheet.cell(row=self.cur_row, column=2)
                cell.value = data["group"]
                self.cur_row += 1
            else:
                if data[str(row - 1)] != "":
                    # self.sheet.merge_cells(start_row=self.cur_row,start_column=1,end_row=self.cur_row+1,end_column=2)
                    cell = self.sheet.cell(row=self.cur_row, column=1)
                    cell.value = str(row - 1) + "\n" + str(self.par_time[row-1]['start'].isoformat("minutes"))
                    cell = self.sheet.cell(row=self.cur_row, column=2)
                    cell.value = data[str(row - 1)]
                    self.cur_row += 1
    def setting_up(self):
        # SHEET CONFIG
        # variables
        center_alignment =Alignment(horizontal='center', vertical='center',wrap_text=True)
        # column settings

        # width
        self.sheet.column_dimensions[get_column_letter(1)].width = 10  # A
        self.sheet.column_dimensions[get_column_letter(2)].width = 35  # B
        # row settings

        # height, alignment
        for row in range(1, self.cur_row):
            self.sheet.row_dimensions[row].height = 50  # height every row
            self.sheet.row_dimensions[1].height = 15 # height first row
            self.sheet.cell(row=row, column=1).alignment = center_alignment
            self.sheet.cell(row=row, column=2).alignment = center_alignment


    def save_file(self,name):
        self.wb.save(name)

    par_time = {
        1: {
            "start": time(hour=8, minute=30),
            "end": time(hour=9, minute=50)
        },
        2: {
            "start": time(hour=10, minute=10),
            "end": time(hour=11, minute=30)
        },
        3: {
            "start": time(hour=11, minute=50),
            "end": time(hour=13, minute=10)
        },
        4: {
            "start": time(hour=13, minute=30),
            "end": time(hour=14, minute=50)
        },
        5: {
            "start": time(hour=15, minute=5),
            "end": time(hour=16, minute=25)
        },
        6: {
            "start": time(hour=16, minute=40),
            "end": time(hour=18, minute=0)
        },
        7: {
            "start": time(hour=18, minute=1),
            "end": time(hour=19, minute=30)
        },
        8: {
            "start": time(hour=19, minute=40),
            "end": time(hour=21, minute=0)
        }
    }


db = DataBase()
# json = {
#     "monday": {
#         "1": "Вишмат",
#         "2": "",
#         "3": "Англійська",
#         "4": "Коман",
#         "5": "",
#         "6": "Бази Даних",
#         "7": "",
#         "8": ""
#     },
#     "wednesday": {
#         "1": "Вишмат",
#         "2": "",
#         "3": "Англійська",
#         "4": "Коман",
#         "5": "",
#         "6": "Бази Даних",
#         "7": "",
#         "8": ""
#     }
# }
# db.set_rozk("Fep22", json)
print(db.get_rozk("Fep22","monday"))
data = {
    "1": "Вишмат",
    "2": "",
    "3": "Англійська",
    "4": "Коман",
    "5": "",
    "6": "Бази Даних",
    "7":"",
    "8":"",
    "group": "ФеП-21",
    "day": "Понеділок"

}

work = Xmlwork(db.get_rozk("Fep22","monday"))
work.save_file("example.xlsx")

excel2img.export_img("example.xlsx","1.png","Первый лист")


# # создаем новый excel-файл
# wb = openpyxl.Workbook()
# # добавляем новый лист
# wb.create_sheet(title='Первый лист', index=0)
#
#
# # получаем лист, с которым будем работать
# sheet = wb['Первый лист']
# print(list(data.keys()))
#
# cur_row = 1
# for row in range(1, 8):
#     if cur_row == 1:
#         cell = sheet.cell(row=cur_row, column=1)
#         cell.value = data["day"]
#         cell = sheet.cell(row=cur_row, column=2)
#         cell.value = data["group"]
#         cur_row +=1
#     else:
#         if data[str(row - 1)] != "":
#             cell = sheet.cell(row=cur_row, column=1)
#             cell.value = "Пара: " + str(row - 1)
#             cell = sheet.cell(row=cur_row, column=2)
#             cell.value = data[str(row - 1)]
#             cur_row +=1
# # SHEET CONFIG
#
# # column settings
# sheet.column_dimensions[get_column_letter(1)].width = 10 # A
# sheet.column_dimensions[get_column_letter(2)].width = 35 # B
# # row settings
# sheet.row_dimensions[1].height = 15
# for row in range(2,cur_row):
#     sheet.row_dimensions[row].height = 50 # height every row
#     sheet.cell(row=row,column=1).alignment = Alignment(horizontal='center',vertical='center')
#     sheet.cell(row=row,column=2).alignment = Alignment(horizontal='center',vertical='center')
#
#
#
# # sheet.row_dimensions['1'].height = 70
# wb.save('example.xlsx')
