import os
from datetime import time


import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Fill, Alignment


class Xlsmwork:

    def __init__(self):
        self.wb = openpyxl.Workbook()


    def fill_full_rozk(self,data):
        # отримує масив даних з розкладом, під кожен день створює лист на якому будуть дані
        for day in data:
            self.wb.create_sheet(title=day['day'],)
            self.sheet = self.wb[day['day']]
            self.doubled_data = []
            self.fill_data(day)
            self.setting_up()
        self.sheet = self.wb['Sheet']
        self.wb.remove_sheet(self.sheet)
        return len(self.wb.sheetnames)

    def fill_data(self, data):
        self.cur_row = 1
        cell = self.sheet.cell(row=self.cur_row, column=1)
        cell.value = data["day"]
        cell = self.sheet.cell(row=self.cur_row, column=2)
        cell.value = data["group"]
        self.cur_row += 1
        data_counter = 1
        for row in range(2, 10):
            if data[str(data_counter)]["doubled"] != None:
                self.doubled_data.append(self.cur_row)


                self.sheet.merge_cells(start_row=self.cur_row, start_column=1, end_row=self.cur_row + 1, end_column=1)
                cell = self.sheet.cell(row=self.cur_row, column=1)
                cell.value = "чис.\n" + str(row - 1) + "\n" + str(self.par_time[row - 1]['start'].isoformat("minutes")) + "\n знам."
                even = data[str(data_counter)]["doubled"]["even"]
                odd = data[str(data_counter)]["doubled"]["odd"]

                cell = self.sheet.cell(row=self.cur_row, column=2)
                cell.value = even
                cell = self.sheet.cell(row=self.cur_row+1, column=2)
                cell.value = odd
                data_counter += 1
                self.cur_row += 2

            elif data[str(data_counter)]["always"] != "":
                self.sheet.merge_cells(start_row=self.cur_row, start_column=2, end_row=self.cur_row + 1, end_column=2)
                self.sheet.merge_cells(start_row=self.cur_row, start_column=1, end_row=self.cur_row + 1, end_column=1)
                cell = self.sheet.cell(row=self.cur_row, column=1)
                cell.value = str(row - 1) + "\n" + str(self.par_time[row - 1]['start'].isoformat("minutes"))
                cell = self.sheet.cell(row=self.cur_row, column=2)
                cell.value = data[str(data_counter)]["always"]

                self.cur_row += 2
                data_counter += 1

    def setting_up(self):
        # SHEET CONFIG
        # variables
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        # column settings

        # width
        self.sheet.column_dimensions[get_column_letter(1)].width = 10  # A
        self.sheet.column_dimensions[get_column_letter(2)].width = 35  # B
        # row settings

        # height, alignment
        self.sheet.row_dimensions[1].height = 15  # height first row
        self.sheet.cell(row=1, column=2).alignment = center_alignment
        self.sheet.cell(row=1, column=1).alignment = center_alignment
        for row in range(2, self.cur_row, 2):
            self.sheet.row_dimensions[row].height = 50  # height every row
            self.sheet.cell(row=row, column=1).alignment = center_alignment
            self.sheet.cell(row=row, column=2).alignment = center_alignment
        for row in self.doubled_data:
            # setting for doubled
            center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            self.sheet.row_dimensions[row].height = 25
            self.sheet.row_dimensions[row + 1].height = 25
            self.sheet.cell(row=row, column=2).alignment = center_alignment
            self.sheet.cell(row=row, column=1).font = Font(size=9)
            self.sheet.cell(row=row + 1, column=2).alignment = center_alignment

            ####

    def save_file(self, filename):
        path = os.path.join(os.path.abspath("tmp"), filename)
        self.wb.save(path)
        return path

    @staticmethod
    def get_sheet_len(path):
        wb =openpyxl.load_workbook(path)
        return wb.sheetnames

    par_time = {
        1: {
            "start": time(hour=8, minute=30),
            "end": time(hour=9, minute=50)
        },
        2: {
            "start": time(hour=10, minute=10),
            "end": time(hour=11, minute=30)
        },
        3: {
            "start": time(hour=11, minute=50),
            "end": time(hour=13, minute=10)
        },
        4: {
            "start": time(hour=13, minute=30),
            "end": time(hour=14, minute=50)
        },
        5: {
            "start": time(hour=15, minute=5),
            "end": time(hour=16, minute=25)
        },
        6: {
            "start": time(hour=16, minute=40),
            "end": time(hour=18, minute=0)
        },
        7: {
            "start": time(hour=18, minute=1),
            "end": time(hour=19, minute=30)
        },
        8: {
            "start": time(hour=19, minute=40),
            "end": time(hour=21, minute=0)
        }
    }


